"""Excel (.xlsx) renderer for proctab Tables.

v0.1: single default theme, single sheet, written to a path on disk via
openpyxl. See docs/EXCEL_RENDERER.md for the locked design memo.

Current state: E1 (format resolver + sheet-name validator) + E2 (column
headers, with merged cells for interior nodes) + E3 (body cells with
MissingReason dispatch and number-format application) + E4 (title at
row 1; source/footnote rows after the body, merged across body cols
with wrap). E5-E7 add default styling for body cells, frozen pane /
widths, and the Table method wiring. openpyxl is an optional extra — install via
`pip install proctab[excel]`. The function-level entry point and the
Table method both lazy-import openpyxl, so `import proctab` works in
environments without it.
"""

from __future__ import annotations

import os
from dataclasses import dataclass

from proctab.model import (
    Category,
    MissingReason,
    Node,
    SubtotalMarker,
    Table,
    TotalMarker,
    ValueKind,
)


# ---------------------------------------------------------------------------
# Format resolution (per docs/EXCEL_RENDERER.md#number-format-resolution).
# ---------------------------------------------------------------------------


_FORMAT_TRANSLATIONS: dict[str, str] = {
    # Integer-like (thousands separator)
    "{:,d}":    "#,##0",
    "{:,.0f}":  "#,##0",
    "{:,.1f}":  "#,##0.0",
    "{:,.2f}":  "#,##0.00",
    "{:,.4f}":  "#,##0.0000",
    # Currency
    "${:,.0f}": "$#,##0",
    "${:,.2f}": "$#,##0.00",
    # Percent — 0-1 scale
    "{:.1%}":   "0.0%",
    "{:.2%}":   "0.00%",
    # Percent — 0-100 scale (literal "%" suffix; freq() flow)
    "{:.1f}%":  '0.0"%"',
    "{:.2f}%":  '0.00"%"',
    # Plain decimals
    "{:.0f}":   "0",
    "{:.1f}":   "0.0",
    "{:.2f}":   "0.00",
    "{:.3f}":   "0.000",
    # General
    "{:g}":     "General",
}
"""Known Python format string → Excel format code translations. Covers
every format `freq()` and `tabulate()` emit today plus the common
hand-written patterns from `examples.py`. Unknown formats fall through
to the per-`value_kind` default."""


_KIND_DEFAULTS: dict[str, str] = {
    "count":         "#,##0",
    "currency":      "$#,##0.00",
    "percent":       "0.0%",            # assumes 0-1 scale
    "ratio":         "0.000",
    "sum":           "#,##0",
    "mean":          "#,##0.00",
    "weighted_mean": "#,##0.00",
    "median":        "#,##0.00",
    "raw":           "General",
}
"""Fallback Excel format codes keyed by `value_kind` (parallel to the
HTML renderer's `_KIND_DEFAULTS`). Used when `formats[j]` is None or
not in the translation table above."""


_GENERAL = "General"


def _resolve_excel_format(fmt: str | None, value_kind: ValueKind) -> str:
    """Resolve the Excel number format code for a single cell.

    Priority (locked in docs/EXCEL_RENDERER.md#number-format-resolution):
      1. Explicit `fmt` translates via `_FORMAT_TRANSLATIONS` when known.
      2. Per-`value_kind` default from `_KIND_DEFAULTS`.
      3. `"General"` final fallback for unknown kinds.
    """
    if fmt is not None and fmt in _FORMAT_TRANSLATIONS:
        return _FORMAT_TRANSLATIONS[fmt]
    return _KIND_DEFAULTS.get(value_kind, _GENERAL)


# ---------------------------------------------------------------------------
# Sheet-name validation (per docs/EXCEL_RENDERER.md, "Frozen pane, column
# widths, sheet name" section).
# ---------------------------------------------------------------------------


_INVALID_SHEET_CHARS = frozenset("\\/?*[]:")
_MAX_SHEET_NAME_LENGTH = 31


def _validate_sheet_name(name: str) -> None:
    """Raise `ValueError` if `name` violates Excel's sheet-name rules.

    Rules: 1-31 characters; must not contain any of `\\ / ? * [ ] :`.
    No silent sanitization — a loud failure beats writing a workbook
    with a corrupted sheet name.
    """
    if not isinstance(name, str):
        raise TypeError(
            f"sheet name must be a string; got {type(name).__name__}."
        )
    if len(name) < 1:
        raise ValueError("sheet name must be at least 1 character.")
    if len(name) > _MAX_SHEET_NAME_LENGTH:
        raise ValueError(
            f"sheet name must be at most {_MAX_SHEET_NAME_LENGTH} "
            f"characters; got {len(name)} ({name!r})."
        )
    bad = sorted(set(name) & _INVALID_SHEET_CHARS)
    if bad:
        raise ValueError(
            f"sheet name contains characters reserved by Excel: {bad}. "
            f"Reserved set: {sorted(_INVALID_SHEET_CHARS)}. "
            f"Rename the sheet or pre-sanitize the string before "
            f"passing it to to_excel()."
        )


# ---------------------------------------------------------------------------
# Layout + shared tree-walking helpers (used by E2 onward).
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class _Layout:
    """Row/column-position variables per docs/EXCEL_RENDERER.md.

    Computed once per `render_excel` call and threaded to helper
    functions so layout math doesn't drift between sections.
    """
    title_present: bool
    H: int                # number of col-axis dim header rows
    n_col_leaves: int
    header_start: int     # first header row (3 if title, else 1)
    body_start: int       # first body row (header_start + H)
    last_body_col: str    # Excel column letter for the last body col


def _nodes_at_depth(root: Node, depth: int) -> list[Node]:
    if root.depth == depth:
        return [root]
    if root.children is None:
        return []
    out: list[Node] = []
    for child in root.children:
        out.extend(_nodes_at_depth(child, depth))
    return out


def _walk_nonroot(node: Node):
    """Pre-order traversal yielding every non-root node (interior + leaf)."""
    if node.depth > 0:
        yield node
    if node.children is not None:
        for child in node.children:
            yield from _walk_nonroot(child)


def _node_label(node: Node) -> str:
    if node.label is not None:
        return node.label
    if not node.path:
        return ""
    el = node.path[-1]
    if isinstance(el, Category):
        return str(el.label if el.label is not None else el.value)
    if isinstance(el, TotalMarker):
        return "Total"
    if isinstance(el, SubtotalMarker):
        return f"{el.at_dim} Subtotal"
    return ""


# ---------------------------------------------------------------------------
# E2 — Column header rendering.
# ---------------------------------------------------------------------------


def _write_thead(ws, col_axis, layout: _Layout) -> None:
    """Write column headers into rows `header_start..header_start + H - 1`.

    Per docs/EXCEL_RENDERER.md, interior nodes that span multiple leaves
    are emitted with merged cells across `node.span` columns; innermost-
    depth leaves are single cells. The top-left corner (column A on
    every header row) is left blank. Per-cell styling lands in E5.
    """
    H = layout.H
    if H == 0:
        # Empty col axis — no headers to emit. Title row stays bare; the
        # body section is also empty.
        return

    for d in range(1, H + 1):
        row = layout.header_start + d - 1
        col = 2  # column A is the blank corner; first content in column B
        for node in _nodes_at_depth(col_axis.tree, d):
            ws.cell(row=row, column=col, value=_node_label(node))
            if node.span > 1:
                ws.merge_cells(
                    start_row=row, start_column=col,
                    end_row=row, end_column=col + node.span - 1,
                )
            col += node.span


# ---------------------------------------------------------------------------
# E3 — Body rendering.
# ---------------------------------------------------------------------------


_MISSING_TEXT: dict[int, str] = {
    int(MissingReason.NOT_APPLICABLE): "—",   # em dash
    int(MissingReason.SUPPRESSED):     "***",
    int(MissingReason.NULL):           "·",   # middle dot
}
"""Display text for non-PRESENT cells (text cells written to Excel).
PRESENT writes a numeric value; EMPTY leaves the cell unwritten."""


def _write_tbody(ws, table: Table, layout: _Layout) -> int:
    """Write body rows starting at `layout.body_start`.

    Pre-orders the row tree: interior nodes emit a single label cell
    at column A (body columns left blank, per the locked memo); leaves
    emit the label at column A plus one cell per col leaf in
    `B..{last_body_col}`.

    Each body cell:
    - Sets `number_format` via `_resolve_excel_format(formats[j],
      value_kinds[j])` so explicit `Table.formats[j]` translations
      (e.g., `freq()`'s `"{:.1f}%"` → `'0.0"%"'`) carry through.
    - Sets `value` per `MissingReason`: PRESENT → `float(body[i,j])`;
      EMPTY → None (cell stays unset); NOT_APPLICABLE / SUPPRESSED /
      NULL → the literal display marker (`—`, `***`, `·`).

    Returns `body_end` — the row number of the last body row written
    (or `body_start - 1` if no rows were emitted), for E4 to anchor
    the source/footnote rows.
    """
    # Empty Table edge case (per docs/EXCEL_RENDERER.md): zero leaves
    # on either axis → no body rows. The col-axis branch matters: a
    # 1×0 Table still has row nodes that would otherwise emit lone
    # column-A labels, which the memo explicitly excludes.
    if layout.n_col_leaves == 0:
        return layout.body_start - 1

    present_code = int(MissingReason.PRESENT)
    empty_code = int(MissingReason.EMPTY)

    row = layout.body_start
    leaf_idx = 0
    body_end = layout.body_start - 1  # if no body rows emitted

    for node in _walk_nonroot(table.row_axis.tree):
        ws.cell(row=row, column=1, value=_node_label(node))

        if node.children is None:
            for j in range(layout.n_col_leaves):
                col = 2 + j
                cell = ws.cell(row=row, column=col)
                cell.number_format = _resolve_excel_format(
                    table.formats[j], table.value_kinds[j]
                )

                missing_code = int(table.missing[leaf_idx, j])
                if missing_code == present_code:
                    cell.value = float(table.body[leaf_idx, j])
                elif missing_code == empty_code:
                    pass  # leave value=None — cell stays blank
                else:
                    cell.value = _MISSING_TEXT[missing_code]

            leaf_idx += 1
        # Interior (group) row: column A label only; B..N left blank.

        body_end = row
        row += 1

    return body_end


# ---------------------------------------------------------------------------
# E4 — Title + source + footnotes.
#
# Title at row 1 (when `meta.title` is set), merged A1:{last_body_col}1,
# bold, font size +2 over default. Spacer row 2 left blank.
# Source row at body_end + 2 (when `meta.source` is set), prefixed
# "Source: ", merged, italic, smaller font, wrap, thin top border.
# Footnote rows from body_end + 3, same merging/wrap/italic/smaller
# font; only the first tfoot row carries the top border (source if
# present, else the first footnote).
# ---------------------------------------------------------------------------


_TITLE_FONT_SIZE = 13  # default 11 + 2
_TFOOT_FONT_SIZE = 10  # default - 1


def _last_col_index(layout: _Layout) -> int:
    """1-based column index of the last body column (or column A when
    the col axis is empty, so title/footer merges stay self-consistent)."""
    return max(1, 1 + layout.n_col_leaves)


def _write_title(ws, table: Table, layout: _Layout) -> None:
    if not table.meta:
        return
    title = table.meta.get("title")
    if not title:
        return

    # openpyxl imports lazily — by the time we get here, render_excel
    # has already done the openpyxl import successfully.
    from openpyxl.styles import Alignment, Font

    cell = ws.cell(row=1, column=1, value=str(title))
    cell.font = Font(bold=True, size=_TITLE_FONT_SIZE)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    end_col = _last_col_index(layout)
    if end_col > 1:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=end_col,
        )


def _write_tfoot(ws, table: Table, layout: _Layout, body_end: int) -> None:
    if not table.meta:
        return
    source = table.meta.get("source")
    footnotes = table.meta.get("footnotes") or []
    if not source and not footnotes:
        return

    from openpyxl.styles import Alignment, Border, Font, Side

    tfoot_font = Font(size=_TFOOT_FONT_SIZE, italic=True)
    wrap_align = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    thin_top_border = Border(top=Side(style="thin"))
    end_col = _last_col_index(layout)

    def _emit(row: int, text: str, *, with_top_border: bool) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = tfoot_font
        cell.alignment = wrap_align
        if with_top_border:
            cell.border = thin_top_border
        if end_col > 1:
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=end_col,
            )

    row = body_end + 2  # blank spacer at body_end + 1
    if source:
        _emit(row, f"Source: {source}", with_top_border=True)
        row += 1

    for i, fn in enumerate(footnotes):
        # Border only on the first footer row overall — source first if
        # it's present, else the first footnote.
        with_top = (not source) and (i == 0)
        _emit(row, str(fn), with_top_border=with_top)
        row += 1


# ---------------------------------------------------------------------------
# Top-level entry point.
# ---------------------------------------------------------------------------


def render_excel(
    table: Table,
    path: str | os.PathLike[str],
    *,
    sheet: str = "Sheet1",
) -> None:
    """Render a `Table` to an `.xlsx` file at `path`.

    Current state: writes a workbook with the column-header band (E2).
    E3-E6 add body, caption/source/footnote, default styling, and the
    frozen pane / column widths.

    `sheet` becomes the worksheet name; must satisfy Excel's rules
    (1-31 characters, no `\\ / ? * [ ] :`). Violations raise
    `ValueError` with a clear message — no silent sanitization.

    Raises `ImportError` with an actionable message if openpyxl is
    not installed (install via `pip install proctab[excel]`).
    """
    # Validate the sheet name BEFORE attempting the openpyxl import so
    # that an invalid sheet= raises the documented ValueError
    # deterministically — even in environments where openpyxl isn't
    # installed (otherwise the ImportError would mask it).
    _validate_sheet_name(sheet)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "Excel export requires openpyxl. Install with "
            "`pip install proctab[excel]`."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    title_present = bool(table.meta.get("title"))
    H = len(table.col_axis.dims)
    n_col_leaves = len(table.col_axis.leaves())
    header_start = 3 if title_present else 1
    body_start = header_start + H
    last_body_col = get_column_letter(max(1, 1 + n_col_leaves))

    layout = _Layout(
        title_present=title_present,
        H=H,
        n_col_leaves=n_col_leaves,
        header_start=header_start,
        body_start=body_start,
        last_body_col=last_body_col,
    )

    _write_title(ws, table, layout)
    _write_thead(ws, table.col_axis, layout)
    body_end = _write_tbody(ws, table, layout)
    _write_tfoot(ws, table, layout, body_end)

    wb.save(path)
